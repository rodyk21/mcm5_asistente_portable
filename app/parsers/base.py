from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.utils import compact_join, normalize_text, slug_text


@dataclass
class ParsedDocument:
    tipo: str
    source_path: Path
    tables: dict[str, list[dict]] = field(default_factory=dict)

    def add(self, table_name: str, row: dict) -> None:
        self.tables.setdefault(table_name, []).append(row)

    @property
    def total_rows(self) -> int:
        return sum(len(rows) for rows in self.tables.values())


def open_workbook(path: Path, *, read_only: bool = True):
    return load_workbook(path, read_only=read_only, keep_vba=False, data_only=True)


def row_values(ws: Worksheet, row_index: int, max_col: int = 60) -> list[object]:
    return [ws.cell(row_index, col).value for col in range(1, max_col + 1)]


def find_row_by_keywords(
    ws: Worksheet,
    start_row: int,
    end_row: int,
    keywords: tuple[str, ...],
    max_col: int = 60,
) -> int | None:
    for row_index in range(start_row, end_row + 1):
        values = [slug_text(value) for value in row_values(ws, row_index, max_col=max_col)]
        if all(any(keyword in value for value in values) for keyword in keywords):
            return row_index
    return None


def row_text(ws: Worksheet, row_index: int, start_col: int = 1, end_col: int = 50) -> str:
    values = [ws.cell(row_index, col).value for col in range(start_col, end_col + 1)]
    return compact_join(values)


def first_text_in_row(
    ws: Worksheet,
    row_index: int,
    min_length: int = 3,
    start_col: int = 1,
    end_col: int = 50,
    skip_tokens: tuple[str, ...] = (),
) -> str | None:
    for col in range(start_col, end_col + 1):
        value = normalize_text(ws.cell(row_index, col).value)
        if len(value) < min_length:
            continue
        if any(token in slug_text(value) for token in skip_tokens):
            continue
        return value
    return None
