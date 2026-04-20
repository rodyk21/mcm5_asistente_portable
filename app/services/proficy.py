from __future__ import annotations

import csv
import io
import re
from pathlib import Path

from openpyxl import load_workbook

from app.utils import normalize_text, parse_float, parse_int, slug_text


def _map_area(location: object) -> str:
    text = slug_text(location)
    if not text:
        return "Sin area"
    if "optima" in text:
        return "Optima"
    if "cuff" in text:
        return "Cuff"
    if "adl" in text:
        return "ADL"
    if "leg" in text:
        return "Leg Die"
    if "flared" in text:
        return "Flared"
    if "outer" in text:
        return "Outer"
    if "ipt" in text:
        return "IPT"
    if "eoc" in text:
        return "EOC"
    return normalize_text(location) or "Sin area"


def _sheet_rows_as_dicts(ws) -> list[dict]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [normalize_text(value) for value in rows[0]]
    data: list[dict] = []
    for raw_row in rows[1:]:
        if not any(value not in (None, "") for value in raw_row):
            continue
        item: dict[str, object] = {}
        for index, header in enumerate(headers):
            if header:
                item[header] = raw_row[index] if index < len(raw_row) else None
        data.append(item)
    return data


def _sheet_rows_as_matrix(ws) -> list[list[object]]:
    return [list(row) for row in ws.iter_rows(values_only=True)]


def _parse_linesummary_downtime(rows: list[dict]) -> list[dict]:
    parsed: list[dict] = []
    for row in rows:
        start_time = normalize_text(row.get("StartTime") or row.get("starttime"))
        end_time = normalize_text(row.get("ENDTime") or row.get("EndTime") or row.get("endtime"))
        duration = parse_float(row.get("Duration") or row.get("duration")) or 0
        fault = normalize_text(row.get("Fault") or row.get("fault"))
        reason1 = normalize_text(row.get("Reason1") or row.get("reason1"))
        reason2 = normalize_text(row.get("Reason2") or row.get("reason2"))
        reason3 = normalize_text(row.get("Reason3") or row.get("reason3"))
        reason4 = normalize_text(row.get("Reason4") or row.get("reason4"))
        location = normalize_text(row.get("Location") or row.get("location"))
        shift = normalize_text(row.get("Shift") or row.get("shift"))
        comments = normalize_text(row.get("Comments") or row.get("comments"))
        status = normalize_text(row.get("Status") or row.get("status"))

        if not start_time or (not fault and duration == 0):
            continue
        if duration == 0 and status == "0":
            continue

        parsed.append(
            {
                "start_time": start_time,
                "end_time": end_time,
                "duracion_min": round(duration),
                "fault": fault,
                "reason1": reason1,
                "reason2": reason2,
                "reason3": reason3,
                "reason4": reason4,
                "area": _map_area(location),
                "location_raw": location,
                "shift": shift,
                "comentario": comments,
                "origen": "proficy_linesummary",
            }
        )
    return parsed


def _find_row_values(rows: list[list[object]], label: str) -> list[str]:
    target = slug_text(label)
    for row in rows:
        cells = [normalize_text(value) for value in row]
        for index, cell in enumerate(cells):
            if target in slug_text(cell):
                return [value for value in cells[index + 1 :] if value]
    return []


def _parse_dpr_summary(rows: list[list[object]], filename: str) -> dict | None:
    fecha = ""
    for row in rows:
        cells = [normalize_text(value) for value in row]
        for index, cell in enumerate(cells):
            if "start date" in slug_text(cell):
                fecha = normalize_text(cells[index + 1] if index + 1 < len(cells) else "")
                break
        if fecha:
            break
    if not fecha:
        match = re.search(r"(\d{4})(\d{2})(\d{2})", filename)
        fecha = f"{match.group(1)}-{match.group(2)}-{match.group(3)}" if match else ""

    line_stops = _find_row_values(rows, "Line Stops")
    downtime = _find_row_values(rows, "Downtime")
    good_product = _find_row_values(rows, "Good Product")
    product = normalize_text((_find_row_values(rows, "Product Code") or _find_row_values(rows, "Product") or [""])[0])
    product = product.replace("All", "").strip()

    return {
        "fecha": fecha,
        "producto": product,
        "paros_t1": parse_int(line_stops[0] if len(line_stops) > 0 else 0) or 0,
        "paros_t2": parse_int(line_stops[1] if len(line_stops) > 1 else 0) or 0,
        "paros_t3": parse_int(line_stops[2] if len(line_stops) > 2 else 0) or 0,
        "paros_total": parse_int(line_stops[3] if len(line_stops) > 3 else (line_stops[2] if len(line_stops) > 2 else 0)) or 0,
        "downtime_t1": parse_float(downtime[0] if len(downtime) > 0 else 0) or 0,
        "downtime_t2": parse_float(downtime[1] if len(downtime) > 1 else 0) or 0,
        "downtime_t3": parse_float(downtime[2] if len(downtime) > 2 else 0) or 0,
        "downtime_total": parse_float(downtime[3] if len(downtime) > 3 else (downtime[2] if len(downtime) > 2 else 0)) or 0,
        "good_product_total": parse_int(good_product[3] if len(good_product) > 3 else (good_product[2] if len(good_product) > 2 else 0)) or 0,
        "origen": "proficy_dpr",
    }


def _read_xlsx(path: Path, filename: str) -> dict:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if "Downtime" in workbook.sheetnames:
            rows = _sheet_rows_as_dicts(workbook["Downtime"])
            return {"type": "paros", "data": _parse_linesummary_downtime(rows)}
        if "Summary" in workbook.sheetnames:
            rows = _sheet_rows_as_matrix(workbook["Summary"])
            parsed = _parse_dpr_summary(rows, filename)
            return {"type": "dpr", "data": [parsed] if parsed else []}

        first_sheet = workbook[workbook.sheetnames[0]]
        rows = _sheet_rows_as_dicts(first_sheet)
        return {"type": "otros", "data": rows}
    finally:
        workbook.close()


def _read_csv_like(path: Path) -> dict:
    text = path.read_text(encoding="utf-8", errors="replace")
    sample = text[:2048]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,")
        reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    except csv.Error:
        delimiter = ";" if sample.splitlines() and sample.splitlines()[0].count(";") >= sample.splitlines()[0].count(",") else ","
        reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)

    rows = [
        {key: normalize_text(value) for key, value in row.items() if key is not None}
        for row in reader
    ]
    headers = " ".join(reader.fieldnames or [])
    if re.search(r"starttime|fault|reason1|duration|downtime", slug_text(headers)):
        return {"type": "paros", "data": _parse_linesummary_downtime(rows)}
    return {"type": "otros", "data": rows}


def parse_proficy_file(path: Path, filename: str) -> dict:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        return _read_xlsx(path, filename)
    if suffix in {".csv", ".txt"}:
        return _read_csv_like(path)
    if suffix == ".xls":
        raise ValueError("Los .xls de Proficy no se pueden leer offline con esta compilacion. Exportalo como .xlsx o .csv y vuelve a importarlo.")
    raise ValueError("Formato Proficy no soportado. Usa .xlsx, .csv o .txt.")
